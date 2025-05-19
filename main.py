import datetime
import openpyxl
workBookName = 'Attendance.xlsx'
Workbook1 = openpyxl.load_workbook(workBookName)
WorkSheet1 = Workbook1['Sheet1']

WorkSheet1["A1"].value = "Name"
WorkSheet1["B1"].value = "Roll No."
WorkSheet1["C1"].value = "Entry-Time"
WorkSheet1["E1"].value = "Sign-up names"
WorkSheet1["F1"].value = "Sign-up Rolls"
WorkSheet1["G1"].value = "Passwords"
WorkSheet1["I1"].value = "Ban Rolls"
WorkSheet1["J1"].value = "Late Students"
WorkSheet1["K1"].value = "Leave Before5"
WorkSheet1["L1"].value = "Leave After5"
WorkSheet1["M1"].value = "Still inCampus"
Workbook1.save(workBookName)


# These first row cell can be comment_out after one time run execution.

# This function help us to to append all data into file to it's given location.
def AppendData(list, startrow, columnnumber):
    for Index in range(startrow, len(list) + startrow):
        rownumber = 1
        while WorkSheet1.cell(row=rownumber, column=columnnumber).value != None:
            rownumber += 1
        WorkSheet1.cell(
            row=rownumber, column=columnnumber).value = list[Index - startrow]
        Workbook1.save(workBookName)
        # Here 'startrow' ask us that from what row number you want to start your appending data,(but it will not overwrite the existing row.)
        # Where 'columnnumber' is a column number.


# This function will help us to read the data from file and append into given list...
def ToRead(list, columnnumber):
    WorkSheet1 = Workbook1['Sheet1']
    Maximumrow = WorkSheet1.max_row  # it will find how much rows are filled.
    for Index in range(2, Maximumrow + 1):
        WorkSheet1 = Workbook1['Sheet1']
        if WorkSheet1.cell(row=Index, column=columnnumber).value != None:
            list.append(WorkSheet1.cell(row=Index, column=columnnumber).value)
    Workbook1.save(workBookName)


def removedata(columnnumber):  # This function will remove all data from the existing given column...Note: But it wil start from the second row to remove data(row number permanently provided into the function as our requirements.)
    WorkSheet1 = Workbook1['Sheet1']
    Maximumrow = WorkSheet1.max_row  # it will find how much rows are filled.
    for Index in range(2, Maximumrow + 1):
        WorkSheet1 = Workbook1['Sheet1']
        Sheet1 = Workbook1.active
        Sheet1[columnnumber + str(Index)] = None
    Workbook1.save(workBookName)


class Attendance:
    Signup_Names = []
    Signup_Rolls = []
    Signup_Passwords = []
    Totalnames = []
    Totalrolls = []
    BanRollNumbers = []
    Entrytime = []
    Lateentry = []
    Leaveafter5 = []
    Leavebefore5 = []
    Stillincampus = []

    # This function would read all data from the file and append into the lists. So it should be run firstly before the execution of the program.
    def ReadAllData(self):
        ToRead(Attendance.Totalnames, 1)
        ToRead(Attendance.Totalrolls, 2)
        ToRead(Attendance.Entrytime, 3)
        ToRead(Attendance.Signup_Names, 5)
        ToRead(Attendance.Signup_Rolls, 6)
        ToRead(Attendance.Signup_Passwords, 7)
        ToRead(Attendance.BanRollNumbers, 9)
        ToRead(Attendance.Lateentry, 10)
        ToRead(Attendance.Leavebefore5, 11)
        ToRead(Attendance.Leaveafter5, 12)
        ToRead(Attendance.Stillincampus, 13)

    # This function goes you to sign-up if you not sign-up first, for attendance, come again here after sign-up ... But if you alraedy sign-up,it will mark your attendance.
    def Entry(self):
        print("Enter your Identity to give attendance  into the Class")
        name = input("Enter your name...\n")
        roll = input("Enter your Roll No. \n")
        password = input("Enter your password. \n")
        self.name = name
        self.roll = roll
        self.__password = password
        if name in Attendance.Signup_Names and roll in Attendance.Signup_Rolls and password in Attendance.Signup_Passwords:
            if roll in Attendance.BanRollNumbers:
                print(
                    "\nSorry! You are banned by the Rector's order.\nYou can't enter into the Institute.")
                print("----------------------------------------------")
            else:
                Attendance.Totalnames = []
                ToRead(Attendance.Totalnames, 1)
                removedata("A")
                Attendance.Totalnames.append(self.name)
                AppendData(Attendance.Totalnames, 1, 1)
                Attendance.Totalrolls = []
                ToRead(Attendance.Totalrolls, 2)
                removedata("B")
                Attendance.Totalrolls.append(self.roll)
                AppendData(Attendance.Totalrolls, 1, 2)
                Attendance.Stillincampus = []
                ToRead(Attendance.Stillincampus, 13)
                removedata("M")
                Attendance.Stillincampus.append(self.name)
                AppendData(Attendance.Stillincampus, 1, 13)
                Presenttime = datetime.datetime.now()
                TimewithMunite = Presenttime.strftime("%H.%M")
                Attendance.Entrytime = []
                ToRead(Attendance.Entrytime, 3)
                removedata("C")
                Attendance.Entrytime.append(TimewithMunite)
                AppendData(Attendance.Entrytime, 1, 3)
                print(
                    "\nOkay! your attendance has been noted with the time", TimewithMunite)
                print("_______________________________________")
                Hour = Presenttime.strftime("%H")
                Hour = int(Hour)
                if Hour > 9:
                    Attendance.Lateentry = []
                    ToRead(Attendance.Lateentry, 10)
                    removedata("J")
                    Attendance.Lateentry.append(self.name)
                    AppendData(Attendance.Lateentry, 1, 10)
        else:
            self.Signup()

    # Be Note... This is only Sign-up, For enter Attendance, click Entry() again...
    def Signup(self):
        print("Please! Sign-up first to become a part of this class.")
        name = input("Enter your name...\n")
        Attendance.Signup_Names = []
        ToRead(Attendance.Signup_Names, 5)
        removedata("E")
        Attendance.Signup_Names.append(name)
        AppendData(Attendance.Signup_Names, 1, 5)
        roll = input("Enter your Roll No. \n")
        Attendance.Signup_Rolls = []
        ToRead(Attendance.Signup_Rolls, 6)
        removedata("F")
        Attendance.Signup_Rolls.append(roll)
        AppendData(Attendance.Signup_Rolls, 1, 6)
        password = input("Enter your password. \n")
        Attendance.Signup_Passwords = []
        ToRead(Attendance.Signup_Passwords, 7)
        removedata("G")
        Attendance.Signup_Passwords.append(password)
        AppendData(Attendance.Signup_Passwords, 1, 7)
        print("\n Great! You signup successfully.")
        print("_______________________________________")

    def Information(self, name):
        if name in Attendance.Totalnames:
            Index = Attendance.Totalnames.index(name)
            print("Name is:", name)
            print("Roll NO. is: ", Attendance.Totalrolls[Index])
            print("Entry time is", Attendance.Entrytime[Index])
            print("___________________________________________")
        else:
            print("You input wrong!\nYour given name does not exist.")

    def Exit(self):
        print("\nYou Clicked exit, Do you want to leave the campus...")
        Leave = input("   Yes: 1    No: Any other key\n")
        if Leave == "1":
            print("Enter your identity to leave the campus.\n")
            Name = input("Enter your name...\n")
            Roll = input("Enter your Roll No. \n")
            Password = input("Enter your password. \n")
            if Name in Attendance.Totalnames and Roll in Attendance.Totalrolls and Password in Attendance.Signup_Passwords:
                PresentTime = datetime.datetime.now()
                TimewithMunite = PresentTime.strftime("%H.%M")
                Hour = PresentTime.strftime("%H")
                Hour = int(Hour)
                Attendance.Stillincampus = []
                ToRead(Attendance.Stillincampus, 13)
                removedata("M")
                Attendance.Stillincampus.remove(Name)
                AppendData(Attendance.Stillincampus, 1, 13)
                if Hour > 17:
                    print("Okay! You leaving time is noted. ", TimewithMunite)
                    Attendance.Leaveafter5 = []
                    ToRead(Attendance.Leaveafter5, 12)
                    removedata("L")
                    Attendance.Leaveafter5.append(Name)
                    AppendData(Attendance.Leaveafter5, 1, 12)
                else:
                    print("Okay! You leaving time is noted. ", TimewithMunite)
                    Attendance.Leavebefore5 = []
                    ToRead(Attendance.Leavebefore5, 11)
                    removedata("K")
                    Attendance.Leavebefore5.append(Name)
                    AppendData(Attendance.Leavebefore5, 1, 11)
            else:
                print("You even get no entry... \nOr may you input wrong !\n ")
        else:
            print("  Okay!")
            print("_______________________________________")

    def banlist(self):
        print("\nWelcome to Banlist...\nOnly Rector can access here.\nYou need to verify by Rector's Password...")
        Namalpassword = input("Enter the password to access Banlist. ")
        if Namalpassword == "Namal123":
            print("This is the list of Ban-Students.\n",
                  Attendance.BanRollNumbers)
            print()
            Edit = input(
                "Do you want to edit this list?\n    Yes:1    No: Any other key.\n")
            if Edit == "1":
                AddOrRemove = input(
                    "What you want to do? Add or remove.\n    Add:1    Remove: Any other key.\n")
                if AddOrRemove == "1":
                    Roll = input("Enter new roll no to ban. \n")
                    Attendance.BanRollNumbers = []
                    ToRead(Attendance.BanRollNumbers, 9)
                    removedata("I")
                    Attendance.BanRollNumbers.append(Roll)
                    AppendData(Attendance.BanRollNumbers, 1, 9)
                else:
                    Roll = input(
                        "Enter the roll no. to remove from the list. \n")
                    if Roll in Attendance.BanRollNumbers:
                        Attendance.BanRollNumbers = []
                        ToRead(Attendance.BanRollNumbers, 9)
                        removedata("I")
                        Attendance.BanRollNumbers.remove(Roll)
                        AppendData(Attendance.BanRollNumbers, 1, 9)
                        print("Your given roll no.", Roll,
                              "removed from the list")
                    else:
                        print("Your given roll no. does not exist into the list")
            else:
                print("Okay. Thanks !")
        else:
            print(
                "You are trying to violate the rules to take access to the Ban list...\n     It's a warning !")


def main():
    a = Attendance()
    a.ReadAllData()  # Load existing data from the Excel file

    while True:
        print("\nCampus Attendance System")
        print("1. Entry (Mark Attendance)")
        print("2. Signup")
        print("3. Get Information")
        print("4. Exit")
        print("5. Ban List (Rector Access)")
        print("0. Quit")

        choice = input("Enter your choice: ")

        if choice == "1":
            a.Entry()
        elif choice == "2":
            a.Signup()
        elif choice == "3":
            name = input("Enter name to get information: ")
            a.Information(name)
        elif choice == "4":
            a.Exit()
        elif choice == "5":
            a.banlist()
        elif choice == "0":
            print("Exiting the system. Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")


if __name__ == "__main__":
    main()
